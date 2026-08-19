import mimetypes
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import get_current_admin, get_current_student, require_admin
from app.models.auth import User
from app.models.people import Admin, Student
from app.models.program import ACTIVE_ENROLLMENT_STATUSES, InternshipProgram, ProgramEnrollment
from app.schemas.people import StudentOut, StudentUpdateRequest
from app.services.activity_log_service import log_activity
from app.services.storage import delete as delete_file, download_response, save
from app.utils.file_validation import read_and_validate_upload

router = APIRouter(prefix="/api/students", tags=["students"])


def _xlsx_date(value) -> str:
    """Export dates as text so Excel never receives a timezone-aware datetime."""
    return value.strftime("%Y-%m-%d") if value else ""


def _student_query(
    db: Session,
    search: str | None = None,
    program_id: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
):
    """Build the admin student list query, including the optional enrollment-date range."""
    query = db.query(Student).options(joinedload(Student.user))
    if search:
        query = query.join(User, User.id == Student.user_id).filter(
            or_(Student.full_name.ilike(f"%{search}%"), User.email.ilike(f"%{search}%"))
        )

    enrollment_filter = db.query(ProgramEnrollment.student_id)
    has_enrollment_filter = False
    if program_id:
        enrollment_filter = enrollment_filter.filter(
            ProgramEnrollment.program_id == program_id,
            ProgramEnrollment.status.in_(ACTIVE_ENROLLMENT_STATUSES),
        )
        has_enrollment_filter = True
    if start_date:
        enrollment_filter = enrollment_filter.filter(ProgramEnrollment.start_date >= start_date)
        has_enrollment_filter = True
    if end_date:
        enrollment_filter = enrollment_filter.filter(ProgramEnrollment.expected_end_date <= end_date)
        has_enrollment_filter = True
    if has_enrollment_filter:
        query = query.filter(Student.id.in_(enrollment_filter))
    return query.order_by(Student.id)


def _student_out(db: Session, student: Student) -> StudentOut:
    out = StudentOut.model_validate(student)
    enrollments = (
        db.query(ProgramEnrollment)
        .filter(ProgramEnrollment.student_id == student.id)
        .order_by(ProgramEnrollment.id.desc())
        .all()
    )
    primary = next((e for e in enrollments if e.status in ACTIVE_ENROLLMENT_STATUSES), enrollments[0] if enrollments else None)
    if primary is not None:
        program = db.get(InternshipProgram, primary.program_id)
        out.enrollment_id = primary.id
        out.program_id = primary.program_id
        out.program_name = program.name if program else None
        out.program_code = program.code if program else None
        out.enrollment_status = primary.status
        out.enrollment_start_date = primary.start_date
        out.enrollment_end_date = primary.expected_end_date
        out.enrollment_suspension_reason = primary.suspension_reason
    return out


@router.get("/me", response_model=StudentOut)
def get_my_profile(student: Student = Depends(get_current_student)):
    return student


@router.put("/me", response_model=StudentOut)
def update_my_profile(
    payload: StudentUpdateRequest,
    student: Student = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    changes = payload.model_dump(exclude_unset=True)
    if student.national_id_verified and any(
        changes.get(field) != getattr(student, field)
        for field in ("national_id_type", "national_id_number")
        if field in changes
    ):
        _ensure_national_id_is_editable(student)
    for field, value in changes.items():
        setattr(student, field, value)
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


NATIONAL_ID_SIDES = ("front", "back")


def _side_or_404(side: str) -> str:
    if side not in NATIONAL_ID_SIDES:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invalid document side")
    return side


def _ensure_national_id_is_editable(student: Student) -> None:
    if student.national_id_verified:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="This student's ID has been verified and is locked. It cannot be changed.",
        )


def _set_national_id_document(student: Student, side: str, file: UploadFile, content: bytes) -> None:
    _ensure_national_id_is_editable(student)
    path_field = f"national_id_document_{side}_path"
    name_field = f"national_id_document_{side}_name"
    existing_path = getattr(student, path_field)
    # Save the replacement first.  Some older documents were stored on the
    # previous server filesystem and no longer exist after switching to Blob;
    # a missing legacy file must never prevent the admin/student from uploading
    # a new valid ID document.
    new_path = save(content, "national_id_documents", file.filename)
    setattr(student, path_field, new_path)
    setattr(student, name_field, file.filename)
    if existing_path and existing_path != new_path:
        try:
            delete_file(existing_path)
        except (HTTPException, RuntimeError):
            # The database now points at the newly saved Blob object. Legacy
            # cleanup is best-effort only and must not roll back the replacement.
            pass


@router.post("/me/national-id-document/{side}", response_model=StudentOut)
async def upload_my_national_id_document(
    side: str,
    file: UploadFile,
    student: Student = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Complete a missing signup ID upload; replacement is an admin-only action."""
    side = _side_or_404(side)
    _ensure_national_id_is_editable(student)
    if getattr(student, f"national_id_document_{side}_path"):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="This ID document has already been submitted. Contact an administrator if it needs to be changed.",
        )
    content, _ext = await read_and_validate_upload(file, settings.MAX_UPLOAD_SIZE_MB)
    _set_national_id_document(student, side, file, content)
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


@router.post("/{student_id}/national-id-document/{side}", response_model=StudentOut)
async def admin_upload_national_id_document(
    student_id: int,
    side: str,
    file: UploadFile,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    side = _side_or_404(side)
    student = db.get(Student, student_id)
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    content, _ext = await read_and_validate_upload(file, settings.MAX_UPLOAD_SIZE_MB)
    _set_national_id_document(student, side, file, content)
    db.add(student)
    log_activity(db, admin.user_id, "admin", f"upload_national_id_document_{side}", "students", student.id, file.filename)
    db.commit()
    db.refresh(student)
    return _student_out(db, student)


@router.post("/{student_id}/national-id/verify", response_model=StudentOut)
def verify_national_id(
    student_id: int,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Approve a student's submitted ID and permanently lock its details and files."""
    student = db.query(Student).options(joinedload(Student.user)).filter(Student.id == student_id).first()
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    if student.national_id_verified:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="This student's ID is already verified and locked")
    if not student.national_id_document_front_path or not student.national_id_document_back_path:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Upload both front and back ID documents before verifying.",
        )
    student.national_id_verified = True
    db.add(student)
    log_activity(db, admin.user_id, "admin", "verify_national_id", "students", student.id, student.full_name)
    db.commit()
    db.refresh(student)
    return _student_out(db, student)


@router.get("/{student_id}/national-id-document/{side}/download")
def download_national_id_document(student_id: int, side: str, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    side = _side_or_404(side)
    student = db.get(Student, student_id)
    doc_path = getattr(student, f"national_id_document_{side}_path", None) if student else None
    if student is None or not doc_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No document uploaded for this student")
    doc_name = getattr(student, f"national_id_document_{side}_name")
    media_type, _ = mimetypes.guess_type(doc_name or doc_path)
    try:
        return download_response(doc_path, filename=doc_name or f"id-document-{side}", media_type=media_type or "application/octet-stream")
    except HTTPException as exc:
        if exc.status_code == status.HTTP_404_NOT_FOUND:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="This ID document file is no longer available in storage. Upload it again using Replace.",
            ) from exc
        raise


@router.get("", response_model=list[StudentOut])
def list_students(
    skip: int = 0,
    limit: int = 50,
    search: str | None = None,
    program_id: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    students = _student_query(db, search, program_id, start_date, end_date).offset(skip).limit(limit).all()
    return [_student_out(db, s) for s in students]


@router.get("/export.xlsx")
def export_students(
    search: str | None = None,
    program_id: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Download the same student set currently selected by the admin's filters."""
    students = _student_query(db, search, program_id, start_date, end_date).all()
    rows = [_student_out(db, student) for student in students]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    headers = [
        "Student ID", "Name", "Email", "Phone", "Date of Birth", "Gender", "Address", "City", "State", "Country",
        "Citizenship", "Institution", "Degree", "Graduation Year", "GitHub", "LinkedIn", "National ID Type",
        "National ID Number", "Program", "Enrollment Status", "Start Date", "End Date", "Joined Date",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append([
            row.id, row.full_name, row.email, row.phone, _xlsx_date(row.dob), row.gender, row.address, row.city, row.state, row.country,
            row.citizenship_status, row.institution, row.degree, row.graduation_year, row.github_url, row.linkedin_url,
            row.national_id_type, row.national_id_number, row.program_name, row.enrollment_status,
            _xlsx_date(row.enrollment_start_date),
            _xlsx_date(row.enrollment_end_date),
            _xlsx_date(row.created_at),
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        max_length = max((len(str(cell.value or "")) for cell in column), default=10)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 12), 36)
    output = BytesIO()
    workbook.save(output)
    filename = "students-export.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{student_id}", response_model=StudentOut)
def get_student(student_id: int, admin=Depends(require_admin), db: Session = Depends(get_db)):
    student = db.query(Student).options(joinedload(Student.user)).filter(Student.id == student_id).first()
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    return _student_out(db, student)


@router.put("/{student_id}", response_model=StudentOut)
def admin_update_student(
    student_id: int,
    payload: StudentUpdateRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Lets an admin correct/complete a student's profile fields — e.g. national ID details —
    reusing the same StudentUpdateRequest shape the student's own self-service PUT /me uses."""
    student = db.query(Student).options(joinedload(Student.user)).filter(Student.id == student_id).first()
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    changes = payload.model_dump(exclude_unset=True)
    if student.national_id_verified and any(
        changes.get(field) != getattr(student, field)
        for field in ("national_id_type", "national_id_number")
        if field in changes
    ):
        _ensure_national_id_is_editable(student)
    for field, value in changes.items():
        setattr(student, field, value)
    db.add(student)
    db.commit()
    db.refresh(student)
    return _student_out(db, student)
